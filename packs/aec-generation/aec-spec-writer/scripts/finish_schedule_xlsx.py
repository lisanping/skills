"""Export a finish schedule (装修做法表) to a styled .xlsx workbook.

Input format (JSON):
[
  {
    "id": "L01",
    "location": "大堂地面",
    "layers": ["20 厚石材面层", "30 厚 1:3 干硬性水泥砂浆", "100 厚 C20 细石混凝土"],
    "thickness_mm": 150,
    "fire_rating": "A",
    "remark": "房间号 101-105"
  },
  ...
]

Usage:
    python finish_schedule_xlsx.py input.json --output finish.xlsx
                                              [--title "标准层装修做法表"]
                                              [--project "杭州办公楼"]
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PREFIX_DESC = {
    "L": "楼地面",
    "EW": "外墙面",
    "W": "内墙面",
    "C": "顶棚",
    "R": "屋面",
    "B": "踢脚",
}

HEADERS = [
    ("编号", 10),
    ("部位 / 房间", 20),
    ("构造做法 (自外向内 / 自上而下)", 60),
    ("厚度合计 (mm)", 14),
    ("防火等级", 10),
    ("备注", 24),
]


def _border() -> Border:
    side = Side(style="thin", color="808080")
    return Border(left=side, right=side, top=side, bottom=side)


def _validate(rows: list[dict]) -> list[str]:
    """Return a list of warning strings; empty list = clean."""
    warnings: list[str] = []
    seen_ids: set[str] = set()
    for i, r in enumerate(rows, start=1):
        rid = r.get("id", "").strip()
        if not rid:
            warnings.append(f"row {i}: missing 'id'")
        elif rid in seen_ids:
            warnings.append(f"row {i}: duplicate id '{rid}'")
        else:
            seen_ids.add(rid)
        prefix = "".join(c for c in rid if c.isalpha())
        if prefix and prefix not in PREFIX_DESC:
            warnings.append(
                f"row {i}: id prefix '{prefix}' not in {sorted(PREFIX_DESC)}"
            )
        if not r.get("layers"):
            warnings.append(f"row {i} ({rid}): missing 'layers'")
        if r.get("fire_rating") and r["fire_rating"] not in {"A", "B1", "B2", "B3"}:
            warnings.append(
                f"row {i} ({rid}): fire_rating '{r['fire_rating']}' "
                f"not one of A/B1/B2/B3 (GB 8624-2012)"
            )
    return warnings


def export(rows: list[dict], output: Path,
           title: str = "装修做法表", project: str = "") -> list[str]:
    warnings = _validate(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "FinishSchedule"

    border = _border()

    # Title block
    ws.cell(row=1, column=1, value=title).font = Font(size=16, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    if project:
        ws.cell(row=2, column=1, value=f"项目:{project}").font = Font(size=11)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))

    header_row = 3 if project else 2

    # Header row
    for col_idx, (label, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 28

    # Data rows
    for r_idx, row in enumerate(rows, start=header_row + 1):
        layers = row.get("layers", [])
        layers_text = "\n".join(f"{i + 1}. {layer}" for i, layer in enumerate(layers))
        values = [
            row.get("id", ""),
            row.get("location", ""),
            layers_text,
            row.get("thickness_mm", ""),
            row.get("fire_rating", ""),
            row.get("remark", ""),
        ]
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
                horizontal="center" if c_idx in (1, 4, 5) else "left",
            )
            cell.border = border
        ws.row_dimensions[r_idx].height = max(20, 16 * max(len(layers), 1))

    # Footer note
    note_row = header_row + 1 + len(rows) + 1
    ws.cell(
        row=note_row,
        column=1,
        value="注:防火等级按《建筑材料及制品燃烧性能分级》GB 8624-2012;"
              "做法层次楼地面/屋面自上而下,墙面/吊顶自外向内。",
    ).font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=len(HEADERS))

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    return warnings


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("input_json", help="JSON file with list of finish rows")
    ap.add_argument("--output", required=True, help="Output .xlsx path")
    ap.add_argument("--title", default="装修做法表")
    ap.add_argument("--project", default="")
    args = ap.parse_args()

    rows = json.loads(Path(args.input_json).read_text(encoding="utf-8"))
    if not isinstance(rows, list):
        print("Input JSON must be a list of objects", file=sys.stderr)
        return 2

    warnings = export(rows, Path(args.output), title=args.title, project=args.project)
    print(f"Wrote {len(rows)} row(s) → {args.output}")
    if warnings:
        print(f"\n{len(warnings)} warning(s):")
        for w in warnings:
            print(f"  - {w}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
